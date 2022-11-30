# this program gets the list of user profile link from an excel sheet, and add their updated marks to the excel sheet itself.
# it is assumed that only one row is used for the heading, and is always present in the excel file.
# close the excel file if it is open in your machine, otherwise, it will not run.

import individual_score as inds
import test_for_score as testscr
import individual_prolems_nos as indpro

from openpyxl import Workbook, load_workbook

wb = load_workbook('TestBook.xlsx') # replace 'TestBook.xlsx' with the path including the file name of the excel file of the student info

ws = wb.active

for row in range(2,ws.max_row+1):
    s = str(ws['A'+str(row)].value) # replace 'A' with the column name containing the gfg profile links. Full links should be added in the excel file for this to work.
    new_score = inds.get_score(s)
    problem_nos = indpro.get_numbers(s)
    if new_score == '_ _':
        ws["B"+str(row)].value = 0 # replace 'B' with the column where you have to write the current total score obtained by the student
    else:
        easy, medium, hard = testscr.get_problem_count(s)
        ws["B"+str(row)].value = int(new_score) # replace 'B' with the column where you have to write the current total score obtained by the student
    if problem_nos == '_ _':
        ws["C"+str(row)].value = 0 # for easy problems
        ws["D"+str(row)].value = 0 # for medium problems
        ws["E"+str(row)].value = 0 # for hard problems
        ws["F"+str(row)].value = 0 # for total from number of problems
    else:
        ws["C"+str(row)].value = int(easy) # for easy problems
        ws["D"+str(row)].value = int(medium) # for medium problems
        ws["E"+str(row)].value = int(hard) # for hard problems
        ws["F"+str(row)].value = int(easy)*2 + int(medium)*4 + int(hard)*8 # for total from number of problems
    wb.save('TestBook.xlsx') # replace 'TestBook.xlsx' with the path including the file name of the excel file of the student info
    
wb.save('TestBook.xlsx') # replace 'TestBook.xlsx' with the path including the file name of the excel file of the student info
